"""
财报模板填充 Skill
将 OCR 提取的结构化数据填入 模板.xlsx
"""
from pathlib import Path
from datetime import datetime
import openpyxl

# 模板公式行（只填数据行，不覆盖公式）
FORMULA_BS = {19, 37, 38, 52, 61, 62, 70, 72, 73, 74, 75}
FORMULA_PL = {20, 23, 25}

from models.bs_alias import BS_ALIAS
from models.pl_alias import PL_ALIAS


def fill_template(customer_name: str, bs_data: dict, pl_data: dict,
                  template_path: Path, output_path: Path) -> Path:
    """
    填模板
    template_path: 源模板文件
    output_path:   输出文件
    """
    wb = openpyxl.load_workbook(template_path)
    tbs = wb['资产负债表']
    tpl = wb['利润表']
    
    # 公司名和日期
    for ws in [tbs, tpl]:
        for c in range(3, 6):
            ws.cell(row=1, column=c).value = customer_name
        for c, y in enumerate([2023, 2024, 2025], 3):
            ws.cell(row=3, column=c).value = datetime(y, 1, 1)
            ws.cell(row=4, column=c).value = datetime(y, 12, 31)
    
    # 填资产负债表
    for y, data in bs_data.items():
        col = {2023: 3, 2024: 4, 2025: 5}[y]
        for r in range(6, 78):
            tn = tbs.cell(row=r, column=2).value
            if tn is None or r in FORMULA_BS: continue
            tn = tn.strip()
            if tn in data:
                tbs.cell(row=r, column=col).value = data[tn]
            elif tn in BS_ALIAS and BS_ALIAS.get(tn) in data:
                tbs.cell(row=r, column=col).value = data[BS_ALIAS[tn]]
    
    # 填利润表
    for y, data in pl_data.items():
        col = {2023: 3, 2024: 4, 2025: 5}[y]
        for r in range(6, 26):
            tn = tpl.cell(row=r, column=2).value
            if tn is None or r in FORMULA_PL: continue
            tn = tn.strip()
            if tn in data:
                tpl.cell(row=r, column=col).value = data[tn]
            elif tn in PL_ALIAS and PL_ALIAS.get(tn) in data:
                tpl.cell(row=r, column=col).value = data[PL_ALIAS[tn]]
    
    wb.save(str(output_path))
    return output_path
